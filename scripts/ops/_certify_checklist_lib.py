"""First-tab all-pending checklist for the USMCA Live Chrome workbook."""

from __future__ import annotations

import json
import re
from datetime import datetime, timezone
from pathlib import Path

from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

CELL_MAX = 32000
OPEN_MARK = "☐ OPEN"
DONE_MARK = "☑ DONE"
SKIP_MARK = "— SKIP / N/A"

FILL_INDEX = PatternFill("solid", fgColor="1E3A5F")
FILL_BLOCK = PatternFill("solid", fgColor="FEE2E2")
FILL_OPEN = PatternFill("solid", fgColor="FEE2E2")
FILL_DONE = PatternFill("solid", fgColor="D1FAE5")
FILL_SKIP = PatternFill("solid", fgColor="E2E8F0")
FONT_W = Font(name="Calibri", bold=True, color="FFFFFF", size=12)
FONT_B = Font(name="Calibri", bold=True, size=10)
FONT_N = Font(name="Calibri", size=9)
THIN = Border(
    left=Side(style="thin", color="CBD5E1"),
    right=Side(style="thin", color="CBD5E1"),
    top=Side(style="thin", color="CBD5E1"),
    bottom=Side(style="thin", color="CBD5E1"),
)
WRAP = Alignment(wrap_text=True, vertical="top")

ID_PREFIX_MODULE = (
    ("ACCT", "accounting"),
    ("BANK", "banking"),
    ("SETL", "settlements"),
    ("SETTL", "settlements"),
    ("FACT", "factoring"),
    ("DISP", "dispatch"),
    ("HOP-ASSIGN", "dispatch"),
    ("BOOK", "dispatch"),
    ("CUST", "customers"),
    ("DRV", "drivers"),
    ("DRIVER", "drivers"),
    ("FLEET", "fleet"),
    ("MAINT", "maintenance"),
    ("WO-", "maintenance"),
    ("SAFETY", "safety"),
    ("INSURANCE", "insurance"),
    ("INS-", "insurance"),
    ("LEGAL", "legal"),
    ("CASH", "cash-flow"),
    ("FUEL", "fuel"),
    ("DOCS", "docs"),
    ("COMPLIANCE", "compliance"),
    ("REPORT", "reports"),
    ("Q8-", "reports"),
    ("LINK-F", "lists"),
    ("LIST", "lists"),
    ("VEND", "vendors"),
    ("INV-", "inventory"),
    ("INVENTORY", "inventory"),
    ("USER", "users"),
    ("TASK", "tasks"),
    ("PROG", "program"),
    ("SYS-", "system"),
    ("CI-", "system"),
    ("MAIN-", "system"),
    ("SCOREBOARD", "program"),
)


def clip(s: str) -> str:
    s = (s or "").replace("\x00", " ").strip()
    if len(s) > CELL_MAX:
        return s[: CELL_MAX - 20] + "\n…[truncated]"
    return s


def infer_module_from_id(fid: str) -> str:
    u = fid.upper()
    for prefix, mod in ID_PREFIX_MODULE:
        if u.startswith(prefix):
            return mod
    return "cross-module"


def extract_finding_id(line: str) -> str | None:
    ext = {".ts", ".tsx", ".mjs", ".sql", ".js", ".md", ".json"}
    for m in re.finditer(r"`([^`]+)`", line):
        s = m.group(1).strip()
        if "/" in s:
            continue
        low = s.lower()
        if any(low.endswith(e) for e in ext):
            continue
        if re.match(r"^[A-Z][A-Z0-9]", s) and len(s) >= 6:
            return s
    return None


def parse_guard_open_rows(guard_board: Path, live_sha: str) -> list[dict]:
    if not guard_board.exists():
        return []
    seen: dict[str, dict] = {}
    for line in guard_board.read_text().splitlines():
        if not line.startswith("|"):
            continue
        head = line[:120].upper()
        if "FIXED CODE" in head:
            continue
        if re.search(r"\bFIXED \(PR", head):
            continue
        if not re.search(r"\bOPEN\b", line):
            continue
        if re.search(r"\bSUPERSEDED\b", line) and "OPEN" not in head:
            continue
        fid = extract_finding_id(line)
        if not fid:
            continue
        detail = line.lstrip("| ").strip()
        m = re.search(r"`" + re.escape(fid) + r"`\s*—\s*(.*)", line)
        if m:
            detail = m.group(1).strip()
        seat = ""
        sm = re.search(
            r"\*\*((?:CC-[123]|Cursor|Codex|Cascade|Devin|OWNER)[^*]{0,80})\*\*",
            line,
        )
        if sm:
            seat = sm.group(1).strip()
        prio = "—"
        if "P0" in line[:200]:
            prio = "P0"
        elif "P1" in line[:200]:
            prio = "P1"
        elif "P2" in line[:200]:
            prio = "P2"
        elif "P3" in line[:200]:
            prio = "P3"
        rec = {
            "id": fid,
            "module": infer_module_from_id(fid),
            "seat": seat or "unassigned",
            "prio": prio,
            "cluster": "GUARD-WORKORDERS OPEN (grep this id on main before coding — board can lag)",
            "detail": clip(detail),
            "done_when": (
                "Root-cause fix + guard that fails on the bug + Live Chrome on current healthz SHA "
                f"({live_sha}) OR named UNVERIFIED blocker. Then mark ☑ DONE. Do not recertify U14."
            ),
            "source": "docs/audit/GUARD-WORKORDERS.md",
            "seq": 5000,
        }
        prev = seen.get(fid)
        if prev is None or len(rec["detail"]) > len(prev["detail"]):
            seen[fid] = rec
    return list(seen.values())


def launch_gate_rows(modules: dict, live_sha: str, main_sha: str) -> list[dict]:
    rows: list[dict] = []
    seq = 10
    gates = [
        ("LAUNCH-FW-01", "all", "Fully-Wired item 1: create→canonical. Every Create/Book writes the same table the list reads (R=W). Empty TMS expected — CREATE labeled TEST."),
        ("LAUNCH-FW-02", "all", "Fully-Wired item 2 / DoD-B: wizard depth — every rendered field is in the submit payload. Nested + Add new writes Lists canonical."),
        ("LAUNCH-FW-03", "accounting", "Fully-Wired item 3: money object when owed (vendor/customer + GL + JE/expense/bill/payment). Flags OFF until owner says turn on. Reuse poster; no new GL math."),
        ("LAUNCH-FW-04", "all", "Fully-Wired item 4: forward linkage nav→route→API→canonical Neon table (never RETIRE), entity-scoped."),
        ("LAUNCH-FW-05", "all", "Fully-Wired item 5: reverse linkage — the related record can find this one (not memo-only)."),
        ("LAUNCH-FW-06", "all", "Fully-Wired item 6: catalogs / entity scope USMCA. Units use owner/lease — NOT operating_company_id."),
        ("LAUNCH-FW-07", "all", "Fully-Wired item 7: surface bar — tabs, search, filter, gear, range, picker, Combobox, modal/drawer, wizard."),
        ("LAUNCH-FW-08", "all", "Fully-Wired item 8: QBO-like chrome (ParityDrawer, calendars, +Create/+Book). QBO sync stays OFF."),
        ("LAUNCH-FW-09", "all", "Fully-Wired item 9: universal picker — + Add new FIRST ROW → Lists creator → appears + selected + survives reload."),
        ("LAUNCH-FW-10", "all", "Fully-Wired item 10: RLS / entity. FORCE RLS. Owner sessions are not a backstop."),
        ("LAUNCH-FW-11", "all", "Fully-Wired item 11: class guard (verify-*.mjs + claimed EVEN verify-step). Guard fails on the bug."),
        ("LAUNCH-FW-12-LIVE-CHROME", "all", f"Fully-Wired item 12 LAST: click live app on healthz version={live_sha}. origin/main {main_sha} may be ahead. CI-green ≠ this box. Live=BLOCKED until walked."),
        ("LAUNCH-ZERO-UNIQUE", "all", "Zero unique leftover (500 / dead click / silent no-op) on the stamp SHA. Board-only '11 modules zero' is not this box."),
        ("LAUNCH-MONEY-57CABBAB", "accounting", "PROGRAM-EXPENSE-DOCUMENT-POSTED-WITHOUT-JE — expense 57cabbab-f06a-4fa3-ad67-877eb2e64b0f status=posted posting_status=unposted (no JE). CC-1 NOW. Reuse poster."),
        ("LAUNCH-INVOICE-AR-JE", "accounting", "INVOICE-SENT-WITHOUT-AR-RECOGNITION-JE — sent invoice open A/R with zero linked JE."),
        ("LAUNCH-HOP-BANK", "banking", "hop.bank — Program matched_invoice_id vs Neon recon on TEST load 065538c8 / L-20260824-0007."),
        ("LAUNCH-SPINE-SILENT-DROP", "accounting", "ACCOUNTING-SPINE-EVENT-FIRE-AND-FORGET-SILENT-DROP — audit trail can miss invoice/bill/payment events."),
        ("LAUNCH-CASH-PULL-500", "cash-flow", "CASH-ENTRIES-CUSTOMER-PARTY-REF-KIND-CHECK-CONSTRAINT-GAP — /cash-flow Manual Daily Projections Pull invoices 500s (CHECK vs party_ref_kind customer)."),
        ("LAUNCH-ACCT-F9408", "accounting", "ACCT-F9408 — cash-forecast Proforma column fake $0 (raw delivery date vs ETA cash date)."),
        ("LAUNCH-CLASS-MUTABLE-SCOPE", "cross-module", "ONE CC-1 sweep: mutable company/record scope on money mutations (safety/maintenance/insurance/fuel/accounting). One helper + one ratcheting guard."),
        ("LAUNCH-CLASS-HIDE-FLAG", "banking", "ONE sweep: hide-flag fail-open — banking + cash-flow + reports. Failed hide-flag read must not include hidden accounts."),
        ("LAUNCH-CLASS-DEACTIVATED-LABEL", "customers", "ONE helper: deactivated-entity labels via RLS join (customer profitability + reports audit subjects)."),
        ("LAUNCH-DOCS-500-FIX", "docs", "DOCS-F6072-REGRESSION-UNIT-EQUIPMENT-500 — ensureLinkEntityExists uses operating_company_id on mdata.units/equipment. Finding PR #16261; FIX still owed (owner/lease + schema-aware guard)."),
        ("LAUNCH-VOID-TEST", "all", "CREATE-TEST-THEN-VOID: after launch, owner voids labeled TEST docs. Empty TMS was expected."),
        ("LAUNCH-DEPLOY-CATCHUP", "system", f"Live healthz {live_sha} vs origin/main {main_sha}. Cursor lead deploys every 5–10 min AND every 5–10 merged PRs, one in-flight. CC never trigger_deploy. Skip PR #15546."),
        ("LAUNCH-SKIP-QBO-TRANSP-TRK", "system", "NOT work until USMCA launched: QBO sync, TRANSP, TRK, TMS→QBO write-back, recertify U14.", True),
        ("LAUNCH-SKIP-FORM425", "form_425", "Do not loop /425c. Not leftover. Mark SKIP when acknowledged.", True),
    ]
    for item in gates:
        fid, module, detail = item[0], item[1], item[2]
        skip = len(item) > 3 and item[3]
        rows.append(
            {
                "id": fid,
                "module": module,
                "seat": "CC-1"
                if any(x in fid for x in ("57CABBAB", "JE", "CLASS", "HOP-BANK", "CASH-PULL", "F9408"))
                else "all seats / Cursor lead",
                "prio": "P0",
                "cluster": "LAUNCH GATE (must be true before USMCA is launched)",
                "detail": clip(detail),
                "done_when": "Proven on live USMCA + current healthz, then ☑ DONE. Skip rows: — SKIP / N/A.",
                "source": "FULLY-WIRED-COMPLETE-BAR + leftover law + session remainder",
                "seq": seq,
                "mark": SKIP_MARK if skip else OPEN_MARK,
            }
        )
        seq += 1

    return rows


def pending_named_rows(pending: list[dict], live_sha: str) -> list[dict]:
    out = []
    for i, p in enumerate(pending):
        out.append(
            {
                "id": p["id"],
                "module": p["module"],
                "seat": p["seat"],
                "prio": "P1",
                "cluster": "NAMED leftover from launch remainder (also on module sheets)",
                "detail": clip(p["note"]),
                "done_when": f"Fixed + guarded + Live Chrome on healthz {live_sha}. Then ☑ DONE.",
                "source": "PENDING[] in build-usmca-live-chrome-certify-xlsx.py",
                "seq": 2000 + i,
            }
        )
    return out


ACCOUNTING_TAB_BUCKET = {
    "Accounting": "Accounting",
    "Bills": "Bills",
    "Expenses": "Expenses",
    "Receipts": "Expenses",
    "Bill Payment": "Bill payment",
    "Invoices": "Invoices",
    "Receive Payment": "Invoices",
    "Collections": "Invoices",
    "Vendors": "Vendors",
    "Customers": "Customers",
    "Reports": "Reports",
}


def display_tab(module: str, raw: str) -> str:
    t = raw or "(untabbed)"
    if module == "accounting":
        return ACCOUNTING_TAB_BUCKET.get(t, "More")
    return t


HOP_LABEL = {
    "picker_law": "picker: + Add new FIRST ROW → Lists creator → selected after reload",
    "qbo_chrome": "chrome: calendar click-through + +Create/+Book + no box-in-box",
    "connectivity": "forward: nav→route→API→canonical table",
    "reverse_link": "reverse: related record finds this row",
    "gl_je": "money: JE / poster (reuse poster; flags OFF until owner)",
    "ap_bill": "money: bill + lines",
    "expense": "money: expense",
    "invoice": "money: invoice + lines",
    "bank": "money: bank match/recon",
    "settlement": "money: settlement",
    "driver": "link: driver",
    "customer": "link: customer",
    "vendor": "link: vendor",
    "unit": "link: unit (owner/lease, not operating_company_id)",
    "trailer": "link: trailer/equipment",
    "load": "link: load",
    "claim": "link: insurance claim",
    "work_order": "link: work order",
    "accident": "link: accident",
    "policy": "link: policy",
    "legal_matter": "link: legal matter",
    "liability": "link: liability/escrow",
    "inventory": "link: inventory/parts",
}


def live_chrome_leaf_rows(modules: dict, live_sha: str) -> list[dict]:
    """One OPEN Live Chrome walk per matrix leaf (module → tab → surface → route)."""
    out = []
    n = 0
    names = ["accounting"] + sorted(m for m in modules if m != "accounting")
    for mod in names:
        for leaf in modules[mod].get("leaves") or []:
            n += 1
            lid = leaf.get("id") or f"leaf-{n}"
            tab = display_tab(mod, leaf.get("tab") or "")
            sub = leaf.get("sub") or ""
            route = leaf.get("route_hint") or ""
            req = leaf.get("required") or []
            hops = []
            for c in req:
                if c.startswith("scenario."):
                    hops.append(f"Program {c}")
                else:
                    hops.append(HOP_LABEL.get(c, c))
            go2310 = []
            if "qbo_chrome" in req:
                go2310.append("GO-2310 calendars must click through (no seize / dead label)")
            if "picker_law" in req:
                go2310.append("GO-2310 nested create = Lists chrome (same wizard, R=W)")
            skip = mod == "form_425"
            detail = (
                f"LIVE CHROME still owed on healthz {live_sha} (older U14/leftover stamps do not count). "
                f"Open USMCA → {mod} → tab {tab} → {sub}. Route {route or '—'}. Leaf `{lid}`.\n"
                f"Walk: {'; '.join(hops) if hops else 'open the surface; hunt 500 / dead click / silent no-op / fake $0'}.\n"
                + (("Also: " + "; ".join(go2310) + ".\n") if go2310 else "")
                + "Empty TMS is expected — CREATE labeled TEST, then this row can be proven. Do not recertify U14."
            )
            out.append(
                {
                    "id": f"LC|{mod}|{lid}",
                    "module": mod,
                    "tab": tab,
                    "surface": sub,
                    "route": route,
                    "leaf": lid,
                    "seat": "assigned URL / GO-2310 hunt",
                    "prio": "P1",
                    "cluster": "LIVE CHROME leaf (item 12) — current healthz only",
                    "detail": clip(detail),
                    "done_when": (
                        f"Clicked on live app at healthz {live_sha}: page opens, create/save/reload if owed, "
                        "calendars/pickers if listed, no 500/dead/silent/fake $0. Named evidence. Then ☑ DONE."
                    ),
                    "source": f"docs/specs/scoreboard/modules/{mod}.required.json",
                    "seq": 1500,
                    "mark": SKIP_MARK if skip else OPEN_MARK,
                }
            )
    # eld has no matrix file
    out.append(
        {
            "id": "LC|eld|stub",
            "module": "eld",
            "tab": "(hidden stub)",
            "surface": "ELD is a hidden stub — not a leftover",
            "route": "—",
            "leaf": "stub",
            "seat": "nobody",
            "prio": "—",
            "cluster": "LIVE CHROME leaf (item 12) — current healthz only",
            "detail": "ELD is a hidden stub. Not missing. Mark SKIP.",
            "done_when": "Acknowledged stub. Mark — SKIP / N/A.",
            "source": "sidebar-config eld hidden",
            "seq": 1500,
            "mark": SKIP_MARK,
        }
    )
    return out


def load_checkoff(path: Path) -> tuple[str, dict[str, str]]:
    if not path.exists():
        return "", {}
    try:
        data = json.loads(path.read_text())
    except json.JSONDecodeError:
        return "", {}
    marks = data.get("marks") or {}
    return str(data.get("live_sha") or ""), {str(k): str(v) for k, v in marks.items()}


def apply_checkoffs(rows: list[dict], marks: dict[str, str], prev_sha: str, live_sha: str) -> None:
    sha_changed = bool(prev_sha) and prev_sha != live_sha
    for rec in rows:
        fid = rec["id"]
        if sha_changed and (
            fid.startswith("LC|") or fid.startswith("LIVE-CHROME") or fid == "LAUNCH-FW-12-LIVE-CHROME"
        ):
            if rec.get("mark") != SKIP_MARK:
                rec["mark"] = OPEN_MARK
            continue
        if fid in marks and rec.get("mark") != SKIP_MARK:
            rec["mark"] = marks[fid]


def save_checkoff(path: Path, live_sha: str, main_sha: str, rows: list[dict]) -> None:
    marks = {
        r["id"]: r.get("mark", OPEN_MARK)
        for r in rows
        if r.get("mark") in {DONE_MARK, SKIP_MARK} and not str(r["id"]).startswith("LAUNCH-SKIP")
    }
    path.write_text(
        json.dumps(
            {
                "updated_at": datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ"),
                "live_sha": live_sha,
                "main_sha": main_sha,
                "row_count": len(rows),
                "marks": marks,
            },
            indent=2,
            sort_keys=False,
        )
        + "\n"
    )


def ingest_xlsx_marks(xlsx: Path) -> dict[str, str]:
    if not xlsx.exists():
        return {}
    from openpyxl import load_workbook

    wb = load_workbook(xlsx, read_only=True, data_only=True)
    try:
        ws = wb[wb.sheetnames[0]]
        marks: dict[str, str] = {}
        for row in ws.iter_rows(min_row=3, max_col=4, values_only=True):
            mark = row[0] if row else None
            fid = row[3] if row and len(row) > 3 else None
            if fid and mark:
                marks[str(fid)] = str(mark)
        return marks
    finally:
        wb.close()


def all_checklist_rows(
    modules: dict,
    pending: list[dict],
    guard_board: Path,
    live_sha: str,
    main_sha: str,
) -> list[dict]:
    merged: dict[str, dict] = {}
    order: list[str] = []

    def add(rec: dict) -> None:
        fid = rec["id"]
        if fid not in merged:
            merged[fid] = rec
            order.append(fid)
            return
        old = merged[fid]
        if len(rec.get("detail") or "") > len(old.get("detail") or ""):
            rec["seq"] = min(old.get("seq", 9999), rec.get("seq", 9999))
            rec["mark"] = old.get("mark", rec.get("mark", OPEN_MARK))
            merged[fid] = rec

    for rec in launch_gate_rows(modules, live_sha, main_sha):
        add(rec)
    for rec in live_chrome_leaf_rows(modules, live_sha):
        add(rec)
    for rec in pending_named_rows(pending, live_sha):
        add(rec)
    for rec in parse_guard_open_rows(guard_board, live_sha):
        add(rec)
    rows = [merged[k] for k in order]
    rows.sort(key=lambda r: (r.get("seq", 9999), r.get("module") or "", r["id"]))
    return rows


def write_master_checklist(
    wb,
    modules: dict,
    pending: list[dict],
    guard_board: Path,
    live_sha: str,
    main_sha: str,
    checkoff_path: Path,
    extra_xlsx: list[Path] | None = None,
) -> int:
    rows = all_checklist_rows(modules, pending, guard_board, live_sha, main_sha)
    prev_sha, marks = load_checkoff(checkoff_path)
    for xp in extra_xlsx or []:
        marks.update(ingest_xlsx_marks(xp))
    apply_checkoffs(rows, marks, prev_sha, live_sha)
    ws = wb.create_sheet("00-ALL-PENDING-CHECKLIST", 0)
    ws.merge_cells("A1:N1")
    open_n = sum(1 for r in rows if r.get("mark", OPEN_MARK) == OPEN_MARK)
    lc_n = sum(1 for r in rows if str(r["id"]).startswith("LC|"))
    banner = (
        f"FIRST TAB — every known missing item + every Live Chrome leaf. "
        f"Live SHA {live_sha} · origin/main {main_sha} · USMCA only. "
        f"Rows {len(rows)} · still {OPEN_MARK}: {open_n} · Live Chrome leaves: {lc_n}. "
        f"Column A dropdown. Filter Cluster to 'LIVE CHROME leaf' for tab-by-tab walks. "
        f"Filter Module for one module. Regen: python3 scripts/ops/build-usmca-live-chrome-certify-xlsx.py "
        f"(checkmarks persist in {checkoff_path.name}; Live Chrome checks RESET if healthz SHA changes). "
        "Not a U14 recertify. CI-green ≠ DONE. Older leftover stamps ≠ this SHA."
    )
    c = ws.cell(1, 1, banner)
    c.fill = FILL_BLOCK
    c.font = Font(name="Calibri", bold=True, size=11, color="7F1D1D")
    c.alignment = Alignment(wrap_text=True, vertical="center")
    ws.row_dimensions[1].height = 56

    headers = [
        "DONE (check here)",
        "#",
        "Cluster",
        "FINDING / task ID",
        "Module",
        "Tab",
        "Surface / sub-tab",
        "Route",
        "Leaf id",
        "Seat",
        "Prio",
        "Every detail (what is missing / what to walk)",
        "Check the box when",
        "Source",
    ]
    for i, h in enumerate(headers, 1):
        cell = ws.cell(2, i, h)
        cell.fill = FILL_INDEX
        cell.font = FONT_W
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        cell.border = THIN
    ws.row_dimensions[2].height = 28

    for n, rec in enumerate(rows, 1):
        r = n + 2
        mark = rec.get("mark", OPEN_MARK)
        vals = [
            mark,
            n,
            rec.get("cluster") or "",
            rec["id"],
            rec.get("module") or "",
            rec.get("tab") or "",
            rec.get("surface") or "",
            rec.get("route") or "",
            rec.get("leaf") or "",
            rec.get("seat") or "",
            rec.get("prio") or "",
            rec.get("detail") or "",
            rec.get("done_when") or "",
            rec.get("source") or "",
        ]
        for col, v in enumerate(vals, 1):
            cell = ws.cell(r, col, v)
            cell.font = FONT_N
            cell.alignment = WRAP
            cell.border = THIN
        ws.cell(r, 1).fill = FILL_SKIP if mark == SKIP_MARK else FILL_OPEN
        ws.cell(r, 1).font = FONT_B
        ws.row_dimensions[r].height = 72 if len(str(vals[11])) > 280 else 40

    last = 2 + len(rows)
    dv = DataValidation(
        type="list",
        formula1=f'"{OPEN_MARK},{DONE_MARK},{SKIP_MARK}"',
        allow_blank=False,
        showDropDown=False,
        showErrorMessage=True,
        errorTitle="Use the dropdown",
        error=f"Pick {OPEN_MARK}, {DONE_MARK}, or {SKIP_MARK}",
    )
    dv.add(f"A3:A{last}")
    ws.add_data_validation(dv)
    ws.conditional_formatting.add(
        f"A3:A{last}",
        FormulaRule(formula=['SEARCH("DONE",$A3)'], fill=FILL_DONE),
    )
    ws.conditional_formatting.add(
        f"A3:A{last}",
        FormulaRule(formula=['SEARCH("OPEN",$A3)'], fill=FILL_OPEN),
    )

    widths = [22, 8, 38, 44, 14, 18, 28, 32, 28, 20, 8, 72, 36, 36]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:N{last}"
    ws.sheet_properties.tabColor = "DC2626"
    save_checkoff(checkoff_path, live_sha, main_sha, rows)
    return len(rows)
