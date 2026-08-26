#!/usr/bin/env python3
"""Build USMCA Live Chrome certify Excel: one sheet per module, tabs as columns."""

from __future__ import annotations

import json
import subprocess
import urllib.request
from collections import defaultdict
from pathlib import Path
import sys

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

sys.path.insert(0, str(Path(__file__).resolve().parent))
from _certify_checklist_lib import write_master_checklist  # noqa: E402

ROOT = Path(__file__).resolve().parents[2]
MODULES_DIR = ROOT / "docs/specs/scoreboard/modules"
GUARD_BOARD = ROOT / "docs/audit/GUARD-WORKORDERS.md"
OUT = ROOT / "docs/lockdown/USMCA-LIVE-CHROME-CERTIFY-INVENTORY-2026-08-26.xlsx"
CHECKOFF = ROOT / "docs/lockdown/usmca-live-chrome-checkoff.json"
HEALTHZ = "https://api.ih35dispatch.com/api/v1/healthz/shallow"


def fetch_live_sha(fallback: str) -> str:
    try:
        with urllib.request.urlopen(HEALTHZ, timeout=12) as resp:
            data = json.loads(resp.read().decode())
            return str(data.get("version") or fallback)
    except Exception:
        return fallback


def fetch_main_sha(fallback: str) -> str:
    try:
        subprocess.run(
            ["git", "-C", str(ROOT), "fetch", "origin", "main", "-q"],
            check=False,
            timeout=30,
        )
        out = subprocess.check_output(
            ["git", "-C", str(ROOT), "rev-parse", "--short", "origin/main"],
            text=True,
            timeout=10,
        )
        return out.strip() or fallback
    except Exception:
        return fallback

LIVE_SHA = "9db9982"
MAIN_SHA = "030935d"
ENTITY = "USMCA 5c854333-6ea5-4faa-af31-67cb272fef80"

COL_TO_CATALOG = {
    "driver": "mdata.drivers",
    "customer": "mdata.customers",
    "vendor": "mdata.vendors",
    "unit": "mdata.units (owner/lease — NOT operating_company_id)",
    "trailer": "mdata.equipment",
    "load": "mdata.loads",
    "ap_bill": "accounting.bills + bill_lines",
    "expense": "accounting.expenses",
    "gl_je": "accounting.journal_entries + postings",
    "invoice": "accounting.invoices + invoice_lines",
    "bank": "banking.bank_transactions / bank_accounts",
    "liability": "driver_finance liabilities / escrow",
    "inventory": "inventory / parts",
    "settlement": "driver_finance.driver_settlements",
    "claim": "insurance.claim",
    "work_order": "maintenance.work_orders",
    "accident": "safety.accident_reports",
    "policy": "insurance.policy",
    "legal_matter": "legal.matters",
}

COL_TO_MODULE = {
    "driver": "Drivers",
    "customer": "Customers",
    "vendor": "Vendors",
    "unit": "Fleet",
    "trailer": "Fleet",
    "load": "Dispatch",
    "ap_bill": "Accounting",
    "expense": "Accounting",
    "gl_je": "Accounting",
    "invoice": "Accounting / Customers",
    "bank": "Banking",
    "liability": "Settlements / Banking escrow",
    "inventory": "Inventory / Maintenance",
    "settlement": "Settlements",
    "claim": "Insurance",
    "work_order": "Maintenance",
    "accident": "Safety",
    "policy": "Insurance",
    "legal_matter": "Legal",
}

CHROME_COLS = {"picker_law", "qbo_chrome", "connectivity", "reverse_link"}
PROCESS_PREFIX = "scenario."

# Accounting locked top-row order (approved PNG + subnav-manifest).
ACCOUNTING_TAB_ORDER = [
    "Accounting",
    "Bills",
    "Expenses",
    "Bill payment",
    "Invoices",
    "Maintenance & shop",
    "Vendors",
    "Customers",
    "Reports",
    "More",
]

# Map matrix `tab` strings onto the approved PNG groups for the Accounting sheet.
ACCOUNTING_TAB_BUCKET = {
    "Accounting": "Accounting",
    "Bills": "Bills",
    "Expenses": "Expenses",
    "Receipts": "Expenses",
    "Bill Payment": "Bill payment",
    "Invoices": "Invoices",
    "Receive Payment": "Invoices",
    "Collections": "Invoices",
    "Vendors": "Vendors",
    "Customers": "Customers",
    "Reports": "Reports",
}

# Pending FINDINGS hung on surfaces (keyword match on tab|sub|route|id). Live=BLOCKED until SHA walk.
PENDING: list[dict] = [
    {"id": "PROGRAM-EXPENSE-DOCUMENT-POSTED-WITHOUT-JE", "module": "accounting", "kw": ["expense", "57cabbab"], "seat": "CC-1", "note": "Expense 57cabbab status=posted posting_status=unposted — no JE. NOW money."},
    {"id": "INVOICE-SENT-WITHOUT-AR-RECOGNITION-JE", "module": "accounting", "kw": ["invoice", "receive payment", "ar aging"], "seat": "CC-1", "note": "Sent invoice open A/R with zero linked JE."},
    {"id": "ACCT-F9408", "module": "accounting", "kw": ["cash forecast", "cash-forecast"], "seat": "CC-1", "note": "Proforma column fake $0 — raw delivery date vs ETA cash date."},
    {"id": "ACCT-MONEY-F6508", "module": "accounting", "kw": ["create", "expense", "bill"], "seat": "CC-1", "note": "Creator draft state not reset on company switch (verify same class as mutable-scope)."},
    {"id": "ACCOUNTING-SPINE-EVENT-FIRE-AND-FORGET-SILENT-DROP", "module": "accounting", "kw": ["audit"], "seat": "CC-1", "note": "Audit trail can miss invoice/bill/payment events."},
    {"id": "TEST-DATA-BANK-MATCH-EXPENSES-DOUBLE-SEEDED-6210", "module": "accounting", "kw": ["expense", "journal"], "seat": "CC-1", "note": "Duplicate VOID-AT-LAUNCH TEST 6210 postings."},
    {"id": "LV-USMCA-FIXED-ASSETS-TRK-BULK-REGISTER", "module": "accounting", "kw": ["fixed asset"], "seat": "CC-1", "note": "USMCA must not expose TRK bulk asset register."},
    {"id": "hop.invoice / hop.gl / hop.revenue", "module": "accounting", "kw": ["invoice", "journal"], "seat": "CC-1", "note": "Program hops — Event-1 JE; invoice#=load# going-forward; balanced JE."},
    {"id": "BANK-ACCOUNT-HIDE-CAPABILITY-FAILURE-FAILS-OPEN", "module": "banking", "kw": ["account", "transaction", "reconcil"], "seat": "CC-1", "note": "Failed hide-flag read includes hidden accounts (also cash-flow, reports). CLASS SWEEP."},
    {"id": "BANK-KPI-FAKE-ZERO-CATCH-CLUSTER", "module": "banking", "kw": ["home", "kpi", "dashboard"], "seat": "CC-1", "note": "KPI catch → fake $0."},
    {"id": "hop.bank", "module": "banking", "kw": ["transaction", "match", "reconcil"], "seat": "CC-1", "note": "Program matched_invoice_id vs Neon recon."},
    {"id": "TEST-DATA-BANK-MATCH-EXPENSES-DOUBLE-SEEDED-6210", "module": "banking", "kw": ["match", "categoriz"], "seat": "CC-1", "note": "Shared TEST 6210 double-seed."},
    {"id": "SETL-F6464", "module": "settlements", "kw": ["deduction", "pending"], "seat": "CC-1", "note": "Failed refetch leaves cached deduction actions live."},
    {"id": "CASH-ADVANCE-OWNER-NOTIFICATION-FAILURE-RETURNS-SUCCESS", "module": "settlements", "kw": ["advance", "cash"], "seat": "CC-1", "note": "Advance commits while owner notify fails closed as success."},
    {"id": "scenario.settlement / pay-run JE", "module": "settlements", "kw": ["pay", "settlement", "close"], "seat": "CC-1", "note": "Prove paid settlement + posted pay-run JE on USMCA."},
    {"id": "HOP-ASSIGN-ZERO-RATECARD-DRIVER-BILLS", "module": "dispatch", "kw": ["assign", "book"], "seat": "CC-1", "note": "0 driver bills from rate card × shortest miles. Codex UI only."},
    {"id": "hop.book Book Load silent", "module": "dispatch", "kw": ["book"], "seat": "Cascade", "note": "Re-verify + Book Load no-op on CURRENT healthz (was 66a7f58)."},
    {"id": "LV-GATEA-DRIVER-EXPIRY-FIXED-FOR-ONE", "module": "dispatch", "kw": ["assign", "hos"], "seat": "CC-3", "note": "Gate A vs 114 drivers missing CDL/medical dates."},
    {"id": "LV-CANCEL-CHARGE-NEVER-BECOMES-AN-INVOICE", "module": "dispatch", "kw": ["cancel"], "seat": "OWNER", "note": "Owner-gated GL decision. $0 live today."},
    {"id": "LV-TRIP-PAIRING-COLLAPSES-N-LEGS-TO-ONE-SIGNAL", "module": "dispatch", "kw": ["trip", "pair"], "seat": "OWNER", "note": "Owner semantics."},
    {"id": "CUST-MONEY-F6105", "module": "customers", "kw": ["payment"], "seat": "CC-1", "note": "Unapply posts to a route that does not exist."},
    {"id": "CUST-MONEY-F6312", "module": "customers", "kw": ["statement", "recurring", "late"], "seat": "CC-1", "note": "Statements/Recurring/Late fees — Live Chrome on current SHA; canonical invoices only."},
    {"id": "CUSTOMER-CREATE-DEAD-CLICK", "module": "customers", "kw": ["create", "hub", "list"], "seat": "Cursor", "note": "Board FIXED retested still intermittent first click."},
    {"id": "CUSTOMER-PROFITABILITY-LABEL-LOST-FOR-DEACTIVATED-CUSTOMERS", "module": "customers", "kw": ["profit", "detail"], "seat": "CC-1", "note": "Deactivated customer revenue shows not-visible. CLASS with reports."},
    {"id": "DRV-MONEY-F6083", "module": "drivers", "kw": ["earning"], "seat": "CC-1", "note": "GET failure paints $0."},
    {"id": "DRV-MONEY-F6106", "module": "drivers", "kw": ["payment method"], "seat": "CC-1", "note": "GET failure paints empty."},
    {"id": "DRV-MONEY-F6110", "module": "driver-hub", "kw": ["hub", "overview", "settle"], "seat": "CC-1", "note": "Hub financial feeds fail-as-zero."},
    {"id": "DRIVER-CREATE-DRUG-SCREEN-ACK-NEVER-PERSISTED", "module": "drivers", "kw": ["create"], "seat": "OWNER", "note": "Owner-gated persist design."},
    {"id": "FLEET-MONEY-F6113", "module": "fleet", "kw": ["trip", "cost", "profit"], "seat": "CC-1", "note": "Trip-cost uses caller company not membership."},
    {"id": "FLEET-MONEY-F6304", "module": "fleet", "kw": ["profile", "finance", "cost"], "seat": "CC-1", "note": "Asset-cost unused SQL param silent null purchase price."},
    {"id": "MAINT-MONEY-F6631", "module": "maintenance", "kw": ["part", "purchase"], "seat": "CC-1", "note": "Parts-purchase mutable company scope. CLASS."},
    {"id": "MAINT-MONEY-F6626", "module": "maintenance", "kw": ["labor", "rate", "wo"], "seat": "CC-1", "note": "WO labor rate window.prompt + mutable scope. CLASS."},
    {"id": "WO-AUTO-BILL-NEVER-POSTS-GL-JE", "module": "maintenance", "kw": ["work order", "complete", "bill"], "seat": "CC-1", "note": "Code claimed FIXED — Live Chrome prove on current SHA."},
    {"id": "SAFETY-MONEY-F6635", "module": "safety", "kw": ["escrow", "forfeit"], "seat": "CC-1", "note": "Escrow forfeiture mutable scope. CLASS."},
    {"id": "SAFETY-MONEY-F6634", "module": "safety", "kw": ["fine"], "seat": "CC-1", "note": "Fine lifecycle mutable scope. CLASS."},
    {"id": "SAFETY-MONEY-F6437", "module": "safety", "kw": ["home", "event"], "seat": "Cursor", "note": "No retry after failed money-bearing read."},
    {"id": "SAFETY-EVENTS-TEST-FIXTURE-LEAK-NO-VOID-MECHANISM", "module": "safety", "kw": ["event", "home"], "seat": "CC-1", "note": "TEST BOX4 rows inflate KPIs; no void column."},
    {"id": "INSURANCE-MONEY-F6628", "module": "insurance", "kw": ["payment", "schedule", "paid"], "seat": "CC-1", "note": "Mark paid mutable company/policy scope. CLASS."},
    {"id": "LEGAL-TEMPLATE-NEW-MODAL-PICKER-LAW-NO-ENTITY-TO-PICK", "module": "legal", "kw": ["template", "create"], "seat": "CC-3", "note": "P3 design-intent; picker with no entity."},
    {"id": "CASH-ENTRIES-CUSTOMER-PARTY-REF-KIND-CHECK-CONSTRAINT-GAP", "module": "cash-flow", "kw": ["manual", "projection", "pull"], "seat": "CC-1", "note": "Pull invoices 500 — CHECK vs party_ref_kind customer."},
    {"id": "BANK-ACCOUNT-HIDE-CAPABILITY-FAILURE-FAILS-OPEN", "module": "cash-flow", "kw": ["prediction", "forecast", "actual"], "seat": "CC-1", "note": "Shared hide-flag fail-open."},
    {"id": "Q8-SUBSCRIPTIONS-DELIVERY-WORKER-MISSING", "module": "reports", "kw": ["saved", "schedul", "email"], "seat": "CC-2", "note": "Scheduled report CRUD; 0/18 emails sent."},
    {"id": "AUDIT-TRAIL-SUBJECT-LABEL-LOST-FOR-DEACTIVATED-ENTITIES", "module": "reports", "kw": ["audit"], "seat": "CC-1", "note": "Deactivated-entity join blanks subjects. CLASS."},
    {"id": "CUSTOMER-PROFITABILITY-LABEL-LOST-FOR-DEACTIVATED-CUSTOMERS", "module": "reports", "kw": ["customer profit"], "seat": "CC-1", "note": "Shared deactivated label class."},
    {"id": "BANK-ACCOUNT-HIDE-CAPABILITY-FAILURE-FAILS-OPEN", "module": "reports", "kw": ["cash flow"], "seat": "CC-1", "note": "Shared hide-flag fail-open."},
    {"id": "FUEL-PLANNER-DASHBOARD-SPEND-QUERY-FAILS-AS-ZERO", "module": "fuel", "kw": ["planner", "dashboard", "home"], "seat": "CC-1", "note": "Spend query failure → authoritative $0."},
    {"id": "FUEL-MONEY-F6535", "module": "fuel", "kw": ["overage", "card"], "seat": "CC-1", "note": "Card-overage window.confirm + mutable scope. CLASS."},
    {"id": "COMPLIANCE-PROPERTY-TAX-RENDITION-RAW-SELECT-NOT-COMBOBOX", "module": "compliance", "kw": ["property", "2290", "tax"], "seat": "CC-3", "note": "P3 raw select vs Combobox."},
    {"id": "DOCS-F6072-REGRESSION-UNIT-EQUIPMENT-500", "module": "docs", "kw": ["unit", "equipment", "upload"], "seat": "Cursor", "note": "ensureLinkEntityExists uses operating_company_id on units/equipment — 42703 500. PR #16261 filed, FIX still owed."},
    {"id": "HOP-ASSIGN-ZERO-RATECARD-DRIVER-BILLS", "module": "program", "kw": ["assign", "hop"], "seat": "CC-1", "note": "Program hop.assign red until rate-card bills exist."},
    {"id": "TMS-SETTLEMENT-AUTO-PAY-CRON (not QBO sync)", "module": "system", "kw": ["job", "cron", "background"], "seat": "CC-1", "note": "Rank BEHIND 57cabbab. QBO jobs out of scope."},
]

FILL_HDR = PatternFill("solid", fgColor="1E3A5F")
FILL_TAB = PatternFill("solid", fgColor="0F766E")
FILL_SURF = PatternFill("solid", fgColor="F0FDFA")
FILL_PEND = PatternFill("solid", fgColor="FEF3C7")
FILL_BLOCK = PatternFill("solid", fgColor="FEE2E2")
FILL_INDEX = PatternFill("solid", fgColor="1E3A5F")
FONT_W = Font(name="Calibri", bold=True, color="FFFFFF", size=12)
FONT_TAB = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
FONT_B = Font(name="Calibri", bold=True, size=10)
FONT_N = Font(name="Calibri", size=9)
FONT_SM = Font(name="Calibri", size=8, italic=True, color="334155")
THIN = Border(
    left=Side(style="thin", color="CBD5E1"),
    right=Side(style="thin", color="CBD5E1"),
    top=Side(style="thin", color="CBD5E1"),
    bottom=Side(style="thin", color="CBD5E1"),
)
WRAP = Alignment(wrap_text=True, vertical="top")


def infer_kind(leaf: dict) -> str:
    blob = f"{leaf.get('id','')} {leaf.get('sub','')}".lower()
    if "wizard" in blob:
        return "Wizard"
    if "paritydrawer" in blob or "drawer" in blob or ".panel." in leaf.get("id", ""):
        return "Drawer / ParityDrawer"
    if "modal" in blob or ".modal." in leaf.get("id", ""):
        return "Modal"
    if "popup" in blob or "popover" in blob or "flyout" in blob:
        return "Popup"
    if "toolbar" in blob or "filter" in blob and "chrome" in blob:
        return "Toolbar / filter"
    if "+ create" in blob or "+ book" in blob or ".create" in leaf.get("id", "") or "create" in blob and "+" in leaf.get("sub", ""):
        return "Create (drawer/wizard)"
    if "+ add" in blob:
        return "Nested + Add new (picker row)"
    if "list" in blob or "hub" in blob or leaf.get("id") == "home":
        return "Page / list"
    return "Page / leaf"


def catalogs_and_links(required: list[str]) -> tuple[str, str, str]:
    cats, mods, chrome = [], [], []
    for c in required or []:
        if c in CHROME_COLS:
            if c == "picker_law":
                chrome.append("+ Add new first row (same Lists creator, R=W)")
            elif c == "qbo_chrome":
                chrome.append("QBO chrome (ParityDrawer / calendar / +Create)")
            elif c == "connectivity":
                chrome.append("Forward: nav→route→API→canonical")
            elif c == "reverse_link":
                chrome.append("Reverse: other module finds this row")
            continue
        if c.startswith(PROCESS_PREFIX):
            chrome.append(f"Program {c}")
            continue
        if c in COL_TO_CATALOG:
            cats.append(COL_TO_CATALOG[c])
        if c in COL_TO_MODULE:
            mods.append(COL_TO_MODULE[c])
    # unique preserve order
    def uniq(xs):
        seen = set()
        out = []
        for x in xs:
            if x not in seen:
                seen.add(x)
                out.append(x)
        return out

    return "\n".join(uniq(cats)) or "—", "\n".join(uniq(mods)) or "—", "\n".join(chrome) or "—"


def match_pending(module: str, leaf: dict) -> str:
    hay = " ".join(
        [
            module,
            leaf.get("id", ""),
            leaf.get("tab", ""),
            leaf.get("sub", ""),
            leaf.get("route_hint", ""),
        ]
    ).lower()
    hits = []
    for p in PENDING:
        if p["module"] != module:
            continue
        if any(k.lower() in hay for k in p["kw"]):
            hits.append(f"{p['id']} [{p['seat']}]\n{p['note']}")
    return "\n\n".join(hits)


def load_modules() -> dict[str, dict]:
    out = {}
    for path in sorted(MODULES_DIR.glob("*.required.json")):
        data = json.loads(path.read_text())
        out[data["module"]] = data
    return out


def display_tab(module: str, raw: str) -> str:
    t = raw or "(untabbed)"
    if module == "accounting":
        return ACCOUNTING_TAB_BUCKET.get(t, "More")
    return t


def tab_order(module: str, leaves: list[dict]) -> list[str]:
    seen = []
    for leaf in leaves:
        t = display_tab(module, leaf.get("tab") or "")
        if t not in seen:
            seen.append(t)
    if module == "accounting":
        extra = [t for t in seen if t not in ACCOUNTING_TAB_ORDER]
        return list(ACCOUNTING_TAB_ORDER) + extra
    return seen


def write_module_sheet(wb: Workbook, module: str, data: dict) -> None:
    leaves = data.get("leaves") or []
    tabs = tab_order(module, leaves)
    by_tab: dict[str, list] = defaultdict(list)
    for leaf in leaves:
        by_tab[display_tab(module, leaf.get("tab") or "")].append(leaf)

    title = module[:31]
    ws = wb.create_sheet(title)
    n_cols = max(len(tabs), 1)

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    cell = ws.cell(1, 1, f"{module.upper()}  ·  Live Chrome CERTIFY  ·  {ENTITY}  ·  healthz {LIVE_SHA}  ·  origin/main {MAIN_SHA}  ·  Live=BLOCKED until item 12 on this SHA")
    cell.fill = FILL_BLOCK
    cell.font = Font(name="Calibri", bold=True, color="7F1D1D", size=12)
    cell.alignment = Alignment(wrap_text=True, vertical="center")
    ws.row_dimensions[1].height = 36

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n_cols)
    cell = ws.cell(
        2,
        1,
        "Columns = product tabs (design / matrix). Down each column: every leaf (list, +Create, drawer, modal, popup). "
        "Then catalogs (canonical tables) + other modules (forward+reverse). Pending FINDINGS on the matching surface. "
        "Empty TMS is expected. QBO sync / TRANSP / TRK = off. GO-2310: calendars + nested create = Lists chrome on every Create/picker.",
    )
    cell.font = FONT_SM
    cell.alignment = WRAP
    ws.row_dimensions[2].height = 32

    labels = ["SURFACE (tab / sub / kind)", "ROUTE", "CATALOGS / CANONICAL TABLES", "LINKED MODULES (F+R)", "CHROME / PICKER / WIRING", "PENDING (Live=BLOCKED)"]

    for col, tab in enumerate(tabs, 1):
        c = ws.cell(3, col, tab)
        c.fill = FILL_TAB
        c.font = FONT_TAB
        c.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
        c.border = THIN
        ws.column_dimensions[get_column_letter(col)].width = 38
    ws.row_dimensions[3].height = 22

    max_leaves = max((len(by_tab[t]) for t in tabs), default=1)
    block = len(labels)  # 6 rows per leaf
    # spacer row 4
    start = 4
    for i in range(max_leaves):
        base = start + i * (block + 1)
        for col, tab in enumerate(tabs, 1):
            items = by_tab[tab]
            if i >= len(items):
                for r in range(block + 1):
                    ws.cell(base + r, col).border = THIN
                continue
            leaf = items[i]
            cats, mods, chrome = catalogs_and_links(leaf.get("required") or [])
            pending = match_pending(module, leaf)
            kind = infer_kind(leaf)
            owned = leaf.get("owned_surface_paths") or leaf.get("surface_path") or ""
            if isinstance(owned, list):
                owned = "; ".join(owned)
            surface = f"{leaf.get('sub') or leaf.get('id')}\n[{kind}]\nid: {leaf.get('id')}"
            if owned:
                surface += f"\nfile: {owned}"
            values = [
                surface,
                leaf.get("route_hint") or "—",
                cats,
                mods,
                chrome,
                pending or "— (unique 500/dead/silent only; HUNT-PASS after walk on this SHA)",
            ]
            fills = [FILL_SURF, FILL_SURF, FILL_SURF, FILL_SURF, FILL_SURF, FILL_PEND if pending else FILL_SURF]
            for r, (lab, val, fl) in enumerate(zip(labels, values, fills)):
                cell = ws.cell(base + r, col, f"{lab}\n{val}")
                cell.fill = fl
                cell.font = FONT_N
                cell.alignment = WRAP
                cell.border = THIN
            ws.cell(base + block, col).border = THIN
        for r in range(block):
            ws.row_dimensions[base + r].height = 72 if r in (0, 2, 3, 4, 5) else 28
        ws.row_dimensions[base + block].height = 8

    ws.freeze_panes = "A4"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True


def write_index(wb: Workbook, modules: dict) -> None:
    ws = wb.create_sheet("01-INDEX-PLAN", 1)
    headers = [
        "Module",
        "Excel tab",
        "Matrix leaves",
        "Named OPEN FINDINGS on this sheet",
        "Seat NOW",
        "Method",
        "Live Chrome (item 12)",
    ]
    for i, h in enumerate(headers, 1):
        c = ws.cell(1, i, h)
        c.fill = FILL_INDEX
        c.font = FONT_W
        c.border = THIN
        c.alignment = Alignment(wrap_text=True)
    ws.row_dimensions[1].height = 22

    method_note = (
        "3 class sweeps (mutable money-scope=CC-1; hide-flag fail-open=CC-1; deactivated-label JOIN=CC-1) "
        "+ unique leftovers serial-by-area WIP≤3. CC-1: 57cabbab FIRST. Not 15-way parallel. Not recertify U14."
    )

    row = 2
    for mod in sorted(modules):
        findings = [p for p in PENDING if p["module"] == mod]
        ids = "\n".join(f"{p['id']} [{p['seat']}]" for p in findings) or "— unique hunt only (GO-2310 calendars + nested create)"
        seats = ", ".join(sorted({p["seat"] for p in findings})) or "assigned URL hunt"
        c0 = ws.cell(row, 1, mod)
        c0.font = FONT_B
        ws.cell(row, 2, mod[:31])
        ws.cell(row, 3, len(modules[mod].get("leaves") or []))
        ws.cell(row, 4, ids)
        ws.cell(row, 5, seats)
        ws.cell(row, 6, method_note if row == 2 else "")
        ws.cell(row, 7, f"BLOCKED until walked on healthz {LIVE_SHA}")
        for col in range(1, 8):
            ws.cell(row, col).alignment = WRAP
            ws.cell(row, col).border = THIN
            if col != 1:
                ws.cell(row, col).font = FONT_N
        ws.row_dimensions[row].height = 48 if findings else 28
        row += 1

    # eld stub
    ws.cell(row, 1, "eld")
    ws.cell(row, 2, "(no matrix file — hidden stub)")
    ws.cell(row, 3, 0)
    ws.cell(row, 4, "NOT leftover — hidden stub")
    ws.cell(row, 5, "—")
    ws.cell(row, 6, "")
    ws.cell(row, 7, "N/A stub")
    for col in range(1, 8):
        ws.cell(row, col).border = THIN
        ws.cell(row, col).alignment = WRAP

    row += 2
    ws.merge_cells(start_row=row, start_column=1, end_row=row + 6, end_column=7)
    ws.cell(
        row,
        1,
        "PLAN (locked method — not a 15th certify)\n"
        "1. CC-1 serial money: PROGRAM-EXPENSE-DOCUMENT-POSTED-WITHOUT-JE (57cabbab) → cash-flow Pull invoices 500 → "
        "ACCT-F9408 Proforma $0 → hop.bank → hide-flag fail-open CLASS → mutable-scope money CLASS → remaining named money IDs.\n"
        "2. Cursor: DOCS-F6072 unit/equipment 500 FIX (owner_company_id / lease) + schema-aware guard. Deploy 5–10. Lead census.\n"
        "3. Standalones serial-by-area, WIP≤3, one PR per area (Rule 27). Auditors: unique FINDING on CURRENT SHA only.\n"
        "4. Launch ≠ this workbook filled. Launch = Fully-Wired 1–12 + zero unique leftover ON live healthz + Program hops with JEs. "
        "Until then: Built progress / Live=BLOCKED.\n"
        "5. QBO sync OFF. TRANSP/TRK OFF. form_425 do-not-loop. CREATE-TEST-THEN-VOID. Never restamp U14.\n"
        f"Sources: docs/specs/scoreboard/modules/*.required.json · accounting subnav-manifest.ts · GUARD-WORKORDERS named OPEN · "
        f"CERTIFIED-MEANS-ZERO-UNIQUE-LEFTOVER · FULLY-WIRED-COMPLETE-BAR. Generated against healthz {LIVE_SHA}.",
    ).alignment = WRAP
    ws.row_dimensions[row].height = 120

    widths = [18, 22, 14, 56, 16, 28, 28]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:G{row - 3}"


def write_register(wb: Workbook, modules: dict) -> None:
    ws = wb.create_sheet("02-FLAT-REGISTER", 2)
    headers = [
        "Module",
        "Tab",
        "Sub / surface",
        "Kind",
        "Leaf id",
        "Route",
        "Catalogs",
        "Linked modules",
        "Chrome / picker / wiring",
        "Pending FINDING",
    ]
    for i, h in enumerate(headers, 1):
        c = ws.cell(1, i, h)
        c.fill = FILL_HDR
        c.font = FONT_W
        c.border = THIN
    r = 2
    for mod in sorted(modules):
        for leaf in modules[mod].get("leaves") or []:
            cats, mods, chrome = catalogs_and_links(leaf.get("required") or [])
            pending = match_pending(mod, leaf)
            vals = [
                mod,
                display_tab(mod, leaf.get("tab") or ""),
                leaf.get("sub") or "",
                infer_kind(leaf),
                leaf.get("id") or "",
                leaf.get("route_hint") or "",
                cats.replace("\n", "; "),
                mods.replace("\n", "; "),
                chrome.replace("\n", "; "),
                pending.replace("\n", " | ") if pending else "",
            ]
            for c, v in enumerate(vals, 1):
                cell = ws.cell(r, c, v)
                cell.font = FONT_N
                cell.alignment = WRAP
                cell.border = THIN
                if pending and c == 10:
                    cell.fill = FILL_PEND
            r += 1
    widths = [14, 22, 36, 18, 32, 36, 40, 32, 40, 50]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:J{r - 1}"


def write_pending_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("03-NAMED-PENDING", 3)
    headers = ["FINDING", "Module", "Seat", "Note", "Cluster"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(1, i, h)
        c.fill = FILL_HDR
        c.font = FONT_W
        c.border = THIN
    cluster = {
        "mutable": "CLASS: mutable company/record scope on money mutations (CC-1 one sweep)",
        "hide": "CLASS: hide-flag fail-open (banking/cash-flow/reports)",
        "deact": "CLASS: deactivated-entity label via RLS join",
        "money": "Money skeleton serial (57cabbab first)",
        "unique": "Unique leftover / owner-gated / Live re-verify",
    }
    def cl(pid: str) -> str:
        p = pid.lower()
        if "663" in p or "6508" in p or "6535" in p or "mutable" in p:
            return cluster["mutable"]
        if "hide" in p or "fail-open" in p or "FAILS-OPEN" in pid:
            return cluster["hide"]
        if "deactiv" in p or "LABEL-LOST" in pid:
            return cluster["deact"]
        if "57cabbab" in p or "WITHOUT-JE" in pid or "hop." in p or "F9408" in pid or "CASH-ENTRIES" in pid:
            return cluster["money"]
        return cluster["unique"]

    seen = set()
    r = 2
    for p in PENDING:
        key = (p["id"], p["module"])
        if key in seen:
            continue
        seen.add(key)
        vals = [p["id"], p["module"], p["seat"], p["note"], cl(p["id"])]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(r, c, v)
            cell.alignment = WRAP
            cell.border = THIN
            cell.font = FONT_N
        ws.row_dimensions[r].height = 36
        r += 1
    for i, w in enumerate([48, 16, 12, 70, 52], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:E{r - 1}"


def main() -> None:
    global LIVE_SHA, MAIN_SHA
    LIVE_SHA = fetch_live_sha(LIVE_SHA)
    MAIN_SHA = fetch_main_sha(MAIN_SHA)
    modules = load_modules()
    wb = Workbook()
    default = wb.active
    wb.remove(default)
    extra = [
        OUT,
        Path.home() / "Desktop" / OUT.name,
        Path.home() / "Desktop/IH35-CURSOR-AUDIT" / OUT.name,
    ]
    n = write_master_checklist(
        wb,
        modules,
        PENDING,
        GUARD_BOARD,
        LIVE_SHA,
        MAIN_SHA,
        CHECKOFF,
        extra_xlsx=extra,
    )
    write_index(wb, modules)
    write_register(wb, modules)
    write_pending_sheet(wb)
    # module sheets: accounting first then alpha
    names = ["accounting"] + sorted(m for m in modules if m != "accounting")
    for name in names:
        write_module_sheet(wb, name, modules[name])
    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    desktop = Path.home() / "Desktop/IH35-CURSOR-AUDIT" / OUT.name
    desktop.parent.mkdir(parents=True, exist_ok=True)
    wb.save(desktop)
    desktop_root = Path.home() / "Desktop" / OUT.name
    wb.save(desktop_root)
    print(f"WROTE {OUT}")
    print(f"WROTE {desktop}")
    print(f"WROTE {desktop_root}")
    print(f"modules={len(modules)} pending_named={len(PENDING)} checklist_rows={n}")


if __name__ == "__main__":
    main()
